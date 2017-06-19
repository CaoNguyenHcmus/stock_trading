from openpyxl import Workbook

#############################################
#read file first

















#############################################
wb = Workbook()

# grab the active worksheet
ws = wb.active
'''
# Data can be assigned directly to cells
ws['A1'] = 42

# Rows can also be appended
ws.append([1, 2, 3])

# Python types will automatically be converted
import datetime
ws['A2'] = datetime.datetime.now()
'''

#STT
#MaCK
#Nganh (?)

# Save the file
wb.save("sample.xlsx")


#############################################
# Idea: parsing 1 ma --> ghi data 1 ma luon
# Update theo nhom nganh


#############################################
