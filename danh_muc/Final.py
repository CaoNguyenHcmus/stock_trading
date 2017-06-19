from openpyxl import Workbook
parameter = ["funcname", "nthreads", "areasize", "threadsize", "testsize", "repeats", "testvol", "testaccess", "time", "bandwidth", "rate"]
wb = Workbook(write_only=True)
# This worksheet will be separate for each functions
ws = wb.create_sheet()
# open file
INPUT = open("M2W_Yoctov170_stats.txt", "r")
#start header for every column
irow = 0;
ws.append(parameter)
irow = irow + 1;
print "====Read one line===="
for line in INPUT:
    list_one_line = []
    #grab the active worksheet
#    wb = wb.active
    print line
#    line = INPUT.readline()
#    print "====Remove header===="
    #header = "RESULT	datetime=1970-01-01 02:41:43	host=localhost	version=1	funcname="
    line = line[line.find('funcname'):]
#    print line

#    print "====Start read data===="
    # Need to while space at the end of test case
    #chuoi = "funcname=ScanWrite32PtrSimpleLoop	nthreads=1	areasize=1024	threadsize=1088	testsize=1088	repeats=5286943	testvol=5752193984	testaccess=1438048496	time=1.4989999999999952252	bandwidth=3837354225.4836678505	rate=1.0423848737852267222e-09 "
    count_equal = line.count('=');
    #print count_equal
    index = 0
    while index < count_equal:
    #    print parameter[index]
        rbegin = line.find('=')
        rend = line.find("\t")
    #    print rend
        #skip string fist character covert to float
        if index == 0:
            list_one_line.append(line[rbegin+1:rend])
        else:
            list_one_line.append(float(line[rbegin+1:rend]))
        field = parameter[index]+ '=' + line[rbegin+1:rend]
#        print field
##        ws['A1'] = line[rbegin+1:rend]
    # Delete first parameter
        line = line.strip(field)
        line = line.strip('\t')
    #    print chuoi
        index = index + 1
#    print "====End read data and start new line===="
#    print "====Begin parse list data in one line to excel===="
#    print list_one_line
    ws.append(list_one_line)
    print irow
    irow = irow + 1
wb.save('M2W_Yoctov170_stats.xlsx')
INPUT.close()
